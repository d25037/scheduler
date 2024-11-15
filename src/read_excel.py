from datetime import datetime

from models import (
    CellError,
    ColumnList,
    DateAndWeekday,
    ExcelRow,
    RowNotFoundError,
    SheetNotFoundError,
    StaffList,
)
from my_logger import my_loguru
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import FilePath


def read_all_schedule(file_path: FilePath) -> Workbook | FileNotFoundError:
    logger = my_loguru()
    try:
        wb: Workbook = load_workbook(file_path, data_only=True)
        logger.info(f"{file_path} have been loaded.")
        logger.debug(f"sheet_names: {wb.sheetnames}")
        return wb
    except FileNotFoundError as e:
        return e


def get_worksheet(
    wb: Workbook, date_and_weekday: DateAndWeekday
) -> SheetNotFoundError | tuple[str, Worksheet]:
    year_str = str(date_and_weekday.date.year) + "年"
    month_str = str(date_and_weekday.date.month) + "月"
    year_month = year_str + month_str

    filtered = list(filter(lambda x: year_month == x, wb.sheetnames))
    if len(filtered) == 0:
        return SheetNotFoundError("Sheetが見つかりませんでした")

    target_sheet = filtered[0]

    return (target_sheet, wb[target_sheet])


def get_excel_row(
    wb: Workbook, date_and_weekday: DateAndWeekday
) -> ExcelRow | SheetNotFoundError | CellError | RowNotFoundError:
    sheet_or_error: SheetNotFoundError | tuple[str, Worksheet] = get_worksheet(
        wb=wb, date_and_weekday=date_and_weekday
    )
    if isinstance(sheet_or_error, SheetNotFoundError):
        return sheet_or_error

    sheet_name, ws = sheet_or_error
    # print(sheet_name)

    for i, search_am_row in enumerate(ws.values):
        day = search_am_row[0]
        if not isinstance(day, datetime):
            continue

        if day.date() == date_and_weekday.date:
            for j, search_pm_row in enumerate(ws.values):
                if j == i + 1:
                    return ExcelRow(
                        weekday=date_and_weekday.weekday,
                        sheet_name=sheet_name,
                        am_row=list(search_am_row),
                        pm_row=list(search_pm_row),
                    )

    return RowNotFoundError("日付に一致する行が見つかりませんでした")


def get_excel_column_number(
    wb: Workbook,
    columns: ColumnList,
    staffs: StaffList,
    date_and_weekday: DateAndWeekday,
):
    sheet_or_error: SheetNotFoundError | tuple[str, Worksheet] = get_worksheet(
        wb=wb, date_and_weekday=date_and_weekday
    )
    logger = my_loguru()
    if isinstance(sheet_or_error, SheetNotFoundError):
        return sheet_or_error

    _sheet_name, ws = sheet_or_error

    row2 = ws[2]

    for i, cell in enumerate(row2):
        try:
            if not isinstance(cell.value, str):
                continue
            staffs.filter_one_by_name(cell.value).col_number = i
        except IndexError:
            try:
                columns.filter_one_by_name(cell.value).col_number = i
            except IndexError:
                continue

    logger.debug(staffs)
    logger.debug(columns)

    return (columns, staffs)


def read_schedule(excel_rows: list[ExcelRow], staffs: StaffList, columns: ColumnList):
    logger = my_loguru()
    for excel_row in excel_rows:
        if excel_row.am_row is None or excel_row.pm_row is None:
            continue

        for column in columns.data:
            if column.col_number is None:
                continue

            am_value = excel_row.am_row[column.col_number]
            pm_value = excel_row.pm_row[column.col_number]

            if isinstance(am_value, str):
                column.schedule[f"{excel_row.weekday}_am"] = am_value
                splitted = am_value.split()
                for item in splitted:
                    item = item.strip()
                    for staff in staffs:
                        if staff.name == item:
                            staff.schedule[f"{excel_row.weekday}_am"] = column.name

            if isinstance(pm_value, str):
                column.schedule[f"{excel_row.weekday}_pm"] = pm_value
                splitted = pm_value.split()
                for item in splitted:
                    item = item.strip()
                    for staff in staffs:
                        if staff.name == item:
                            staff.schedule[f"{excel_row.weekday}_pm"] = column.name

        for staff in staffs:
            am_value = excel_row.am_row[staff.col_number]
            pm_value = excel_row.pm_row[staff.col_number]

            if isinstance(am_value, str):
                staff.schedule[f"{excel_row.weekday}_am"] = am_value

            if isinstance(pm_value, str):
                staff.schedule[f"{excel_row.weekday}_pm"] = pm_value

    logger.debug(columns)
    logger.debug(staffs)

    return (columns, staffs)
