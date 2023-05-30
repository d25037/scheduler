from logging import Logger

from models import AppSettings, ColumnList, FilePath, StaffList
from openpyxl import Workbook, load_workbook


def read_template(file_path: FilePath, logger: Logger) -> Workbook | FileNotFoundError:
    try:
        logger = logger.getChild(__name__)
        wb: Workbook = load_workbook(file_path)
        logger.info(f"{file_path} have been loaded.")
        return wb
    except FileNotFoundError as e:
        return e


def write_cells(
    wb: Workbook,
    staffs: StaffList,
    columns: ColumnList,
    noon_room: dict[str, dict[str, str]],
    settings: AppSettings,
):
    ws = wb["py"]

    days = [
        "monday_am",
        "monday_pm",
        "tuesday_am",
        "tuesday_pm",
        "wednesday_am",
        "wednesday_pm",
        "thursday_am",
        "thursday_pm",
        "friday_am",
        "friday_pm",
    ]

    for i in range(100):
        row = 2 + i

        col_a = ws.cell(row=row, column=1).value
        if not isinstance(col_a, str):
            continue

        match col_a:
            case "日付":
                ws.cell(
                    row=row, column=2, value=settings.monday().date.strftime("%Y%m%d")
                )
                ws.cell(
                    row=row, column=3, value=settings.tuesday().date.strftime("%Y%m%d")
                )
                ws.cell(
                    row=row,
                    column=4,
                    value=settings.wednesday().date.strftime("%Y%m%d"),
                )
                ws.cell(
                    row=row, column=5, value=settings.thursday().date.strftime("%Y%m%d")
                )
                ws.cell(
                    row=row, column=6, value=settings.friday().date.strftime("%Y%m%d")
                )
            case "昼CT":
                ws.cell(row=row, column=2, value=noon_room["CT"].get("monday"))
                ws.cell(row=row, column=3, value=noon_room["CT"].get("tuesday"))
                ws.cell(row=row, column=4, value=noon_room["CT"].get("wednesday"))
                ws.cell(row=row, column=5, value=noon_room["CT"].get("thursday"))
                ws.cell(row=row, column=6, value=noon_room["CT"].get("friday"))

            case "昼MR":
                ws.cell(row=row, column=2, value=noon_room["MR"].get("monday"))
                ws.cell(row=row, column=3, value=noon_room["MR"].get("tuesday"))
                ws.cell(row=row, column=4, value=noon_room["MR"].get("wednesday"))
                ws.cell(row=row, column=5, value=noon_room["MR"].get("thursday"))
                ws.cell(row=row, column=6, value=noon_room["MR"].get("friday"))

            case "CT" | "MR" | "RI" | "IVR" | "外勤" | "センター" | "みなとみらい":
                for num, value in enumerate(days):
                    member = staffs.filter_by_schedule(schedule=value, work=col_a)
                    if len(member) == 0:
                        continue
                    for j, mem in enumerate(member):
                        ws.cell(row=row + j, column=2 + num, value=mem.name)

            case "その他":
                exclude_words = [
                    "CT",
                    "MR",
                    "RI",
                    "IVR",
                    "①",
                    "②",
                    "③",
                    "外勤",
                    "センター",
                    "みなとみらい",
                    "US",
                    "藤沢",
                ]

                for num, value in enumerate(days):
                    member = staffs.filter_by_schedule_exclude_list(
                        schedule=value, excludes=exclude_words
                    )

                    if len(member) == 0:
                        continue
                    for j, mem in enumerate(member):
                        ws.cell(row=row + j, column=2 + num, value=mem.name)

            case "休暇":
                for num, value in enumerate(days):
                    member = staffs.filter_by_schedule_list(
                        schedule=value, works=["①", "②", "③"]
                    )
                    if len(member) == 0:
                        continue
                    for j, mem in enumerate(member):
                        ws.cell(row=row + j, column=2 + num, value=mem.name)

            case "藤沢保険":
                for num, value in enumerate(days):
                    member = staffs.filter_by_schedule(schedule=value, work="藤沢")
                    if len(member) == 0:
                        continue
                    for j, mem in enumerate(member):
                        ws.cell(row=row + j, column=2 + num, value=mem.name)

            case "超音波":
                for num, value in enumerate(days):
                    member = staffs.filter_by_schedule(schedule=value, work="US")
                    if len(member) == 0:
                        continue
                    for j, mem in enumerate(member):
                        ws.cell(row=row + j, column=2 + num, value=mem.name)

            case "非常勤":
                part_time = columns.filter_one_by_name(value="非常勤")
                if len(part_time.schedule) == 0:
                    continue

                for num, value in enumerate(days):
                    name = part_time.schedule.get(value)
                    if name is None:
                        continue

                    ws.cell(row=row, column=2 + num, value=name)

            case "イベント":
                free_writing = columns.filter_one_by_name(value="自由記載欄")
                meeting = columns.filter_one_by_name(value="医局会")
                if len(meeting.schedule) == 0 and len(free_writing.schedule) == 0:
                    continue

                for num, value in enumerate(days):
                    free_value = free_writing.schedule.get(value)
                    meeting_value = meeting.schedule.get(value)

                    if isinstance(free_value, str):
                        ws.cell(row=row, column=2 + num, value=free_value)

                    if isinstance(meeting_value, str):
                        ws.cell(row=row, column=2 + num, value="医局会")

            case "初期研修医":
                break

            case _:
                continue

    return wb
